from __future__ import annotations

import argparse
import json
import math
import os
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
import requests
from openpyxl import Workbook

FMP_BASE_URL = "https://financialmodelingprep.com/stable"
REPO_ROOT = Path(__file__).resolve().parents[1]
SECTOR_DIR = REPO_ROOT / "sector"
MASTER_JSON_PATH = REPO_ROOT / "Ticker-Master.json"
DAILY_WORKBOOK_PREFIX = "investment_data"
DATE_FORMAT = "%m%d%Y"
TICKER_REVIEW_PREFIX = "ticker_review"

PILLAR_TARGET_PE_MULTIPLES: Dict[str, int] = {
    "ai_power_generation": 22,
    "ai_networking": 40,
    "ai_memory_and_storage": 18,
    "ai_land": 35,
    "ai_heating_and_cooling": 28,
    "ai_grid_and_power_transmission": 25,
    "ai_electrical_infrastructure": 25,
    "ai_construction": 22,
    "ai_cybersecurity": 55,
    "ai_semiconductor_chip_manufacturing": 22,
    "ai_server_and_rack_systems": 18,
    "ai_software_and_platforms": 65,
    "ai_backup_power": 30,
    "ai_cloud": 30,
    "ai_fire_detection": 22,
    "ai_water": 25,
    "ai_insurance_and_risk": 25,
    "ai_security": 22,
    "ai_compute_chips": 45,
    "default": 25,
}

KNOWN_SHEET_NAMES = {
    "AI Power Generation": "AI Power Gen",
    "AI Networking": "AI Networking",
    "AI Memory and Storage": "AI Memory Storage",
    "AI Land": "AI Land",
    "AI Heating and Cooling": "AI Heat Cooling",
    "AI Grid and Power Transmission": "AI Grid Transmission",
    "AI Electrical Infrastructure": "AI Electrical Infra",
    "AI Construction": "AI Construction",
    "AI Cybersecurity": "AI Cybersecurity",
    "AI Semiconductor Chip Manufacturing": "AI Semi Mfg",
    "AI Server and Rack Systems": "AI Server Rack",
    "AI Software and Platforms": "AI Software Platform",
    "AI Backup Power": "AI Backup Power",
    "AI Cloud": "AI Cloud",
    "AI Fire Detection": "AI Fire Detection",
    "AI Water": "AI Water",
    "AI Insurance and Risk": "AI Insurance Risk",
    "AI Security": "AI Security",
    "AI Compute Chips": "AI Compute Chips",
}

COLUMNS = [
    "Ticker",
    "Current Price",
    "RSI",
    "P/E Ratio",
    "P/S Ratio",
    "Market Cap",
    "Buy Zone",
    "Target Price",
    "Graham Undervalued",
    "Upside %",
    "Revenue Growth %",
    "EPS Growth %",
    "Backlog Growth %",
    "Gross Margin %",
    "Net Cash Ratio",
    "Investment Score",
]


@dataclass(frozen=True)
class TickerContext:
    sector: str
    pillar: str
    tickers: List[str]


class FMPClient:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.session = requests.Session()
        self._ratios_bulk_cache: Optional[Dict[str, dict]] = None

    def _get(self, endpoint: str, symbol: str, symbol_param: str = "symbol") -> Optional[dict]:
        params = {symbol_param: symbol, "apikey": self.api_key}
        url = f"{FMP_BASE_URL}/{endpoint}"
        response = self.session.get(url, params=params, timeout=10)
        if not response.ok:
            print(f"[WARN] FMP {endpoint} returned {response.status_code} for {symbol}: {response.text[:200]}")
            return None
        try:
            payload = response.json()
        except Exception as exc:
            print(f"[WARN] FMP {endpoint} returned invalid JSON for {symbol}: {exc}")
            return None
        if isinstance(payload, list) and payload:
            return payload[0]
        if isinstance(payload, dict):
            return payload
        return None

    def _get_rsi(self, symbol: str) -> Optional[float]:
        url = f"{FMP_BASE_URL}/technical-indicators/rsi"
        response = self.session.get(
            url,
            params={
                "symbol": symbol,
                "periodLength": 10,
                "timeframe": "1day",
                "apikey": self.api_key,
            },
            timeout=10,
        )
        if not response.ok:
            print(f"[WARN] FMP technical-indicators/rsi returned {response.status_code} for {symbol}: {response.text[:200]}")
            return None
        try:
            payload = response.json()
        except Exception as exc:
            print(f"[WARN] FMP technical-indicators/rsi returned invalid JSON for {symbol}: {exc}")
            return None
        if isinstance(payload, list) and payload:
            return self._to_float(payload[0].get("rsi"))
        return None

    def _to_float(self, value: object) -> Optional[float]:
        if value in (None, ""):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _get_ratios_ttm_bulk(self) -> Dict[str, dict]:
        if self._ratios_bulk_cache is not None:
            return self._ratios_bulk_cache

        url = f"{FMP_BASE_URL}/ratios-ttm-bulk"
        for attempt in range(3):
            response = self.session.get(url, params={"apikey": self.api_key}, timeout=60)
            if response.status_code == 429:
                wait = 60 * (attempt + 1)
                print(f"[WARN] FMP ratios-ttm-bulk rate-limited (429), retrying in {wait}s...")
                time.sleep(wait)
                continue
            if not response.ok:
                print(f"[WARN] FMP ratios-ttm-bulk returned {response.status_code}: {response.text[:200]}")
                self._ratios_bulk_cache = {}
                return {}
            break
        else:
            print("[WARN] FMP ratios-ttm-bulk failed after 3 attempts — P/E, P/S, Gross Margin and TTM EPS will be empty")
            self._ratios_bulk_cache = {}
            return {}
        try:
            payload = response.json()
        except Exception as exc:
            print(f"[WARN] FMP ratios-ttm-bulk returned invalid JSON: {exc}")
            self._ratios_bulk_cache = {}
            return {}

        cache: Dict[str, dict] = {}
        if isinstance(payload, list):
            for row in payload:
                if not isinstance(row, dict):
                    continue
                symbol = str(row.get("symbol", "")).strip().upper()
                if symbol:
                    cache[symbol] = row

        self._ratios_bulk_cache = cache
        return cache

    def _get_forward_eps(self, symbol: str) -> Optional[float]:
        # Primary: annual analyst estimates — field is epsAvg
        analyst_estimates_available = False
        try:
            response = self.session.get(
                f"{FMP_BASE_URL}/analyst-estimates",
                params={"symbol": symbol, "period": "annual", "page": 0, "limit": 10, "apikey": self.api_key},
                timeout=10,
            )
            response.raise_for_status()
            payload = response.json()
            if isinstance(payload, list) and payload:
                analyst_estimates_available = True
                for row in payload:
                    if not isinstance(row, dict):
                        continue
                    val = self._to_float(row.get("epsAvg"))
                    if val is not None:
                        return val
        except Exception:
            pass

        # Fallback: earnings endpoint — only if analyst-estimates returned no data at all
        if analyst_estimates_available:
            return None

        try:
            response = self.session.get(
                f"{FMP_BASE_URL}/earnings",
                params={"symbol": symbol, "apikey": self.api_key},
                timeout=10,
            )
            response.raise_for_status()
            payload = response.json()
            if isinstance(payload, list):
                for row in payload:
                    if not isinstance(row, dict):
                        continue
                    if row.get("epsActual") is not None:
                        continue  # already reported — skip
                    val = self._to_float(row.get("epsEstimated"))
                    if val is not None:
                        return val
        except Exception:
            pass

        return None

    def _get_book_value_per_share(self, symbol: str) -> Optional[float]:
        url = f"{FMP_BASE_URL}/ratios"
        response = self.session.get(url, params={"symbol": symbol, "apikey": self.api_key}, timeout=10)
        if not response.ok:
            print(f"[WARN] FMP ratios returned {response.status_code} for {symbol}: {response.text[:200]}")
            return None
        try:
            payload = response.json()
        except Exception as exc:
            print(f"[WARN] FMP ratios returned invalid JSON for {symbol}: {exc}")
            return None
        if isinstance(payload, list) and payload:
            return self._to_float(payload[0].get("bookValuePerShare"))
        return None

    def get_metrics(self, ticker: str) -> Dict[str, Optional[float]]:
        # Use FMP batch-quote endpoint for current price/market-cap parity with API docs.
        quote = self._get("batch-quote", ticker, symbol_param="symbols") or {}
        rsi = self._get_rsi(ticker)
        market_cap_data = self._get("market-capitalization-batch", ticker, symbol_param="symbols") or {}
        ratios = self._get_ratios_ttm_bulk().get(ticker.upper(), {})
        growth = self._get("financial-growth", ticker) or {}
        balance = self._get("balance-sheet-statement", ticker) or {}
        book_value_per_share = self._get_book_value_per_share(ticker)
        forward_eps = self._get_forward_eps(ticker)

        market_cap = self._to_float(market_cap_data.get("marketCap")) or self._to_float(quote.get("marketCap"))
        cash_and_cash_equivalents = self._to_float(balance.get("cashAndCashEquivalents"))
        total_debt = self._to_float(balance.get("totalDebt"))
        net_debt = self._to_float(balance.get("netDebt"))
        if net_debt is None and total_debt is not None and cash_and_cash_equivalents is not None:
            net_debt = total_debt - cash_and_cash_equivalents
        revenue_growth = self._to_float(growth.get("revenueGrowth"))
        eps_growth = self._to_float(growth.get("epsgrowth"))
        if eps_growth is None:
            eps_growth = self._to_float(growth.get("epsdilutedGrowth"))
        gross_margin = self._to_float(ratios.get("grossProfitMarginTTM"))
        if gross_margin is None:
            # Only call income-statement when ratios-ttm-bulk didn't supply gross margin
            income = self._get("income-statement", ticker) or {}
            gross_margin = self._to_float(income.get("grossProfitRatio"))
        ttm_eps = self._to_float(ratios.get("epsTTM"))

        return {
            "Current Price": quote.get("price"),
            "RSI": rsi,
            "P/E Ratio": self._to_float(ratios.get("priceToEarningsRatioTTM")),
            "P/S Ratio": self._to_float(ratios.get("priceToSalesRatioTTM")) or self._to_float(ratios.get("priceToSalesRatio")),
            "Market Cap": market_cap,
            "Forward EPS": forward_eps,
            "TTM EPS": ttm_eps,
            "Book Value Per Share": book_value_per_share,
            "Revenue Growth %": revenue_growth * 100 if revenue_growth is not None else None,
            "EPS Growth %": eps_growth * 100 if eps_growth is not None else None,
            "Gross Margin %": gross_margin * 100 if gross_margin is not None else None,
            "Net Cash Ratio": (
                (-net_debt) / market_cap
                if net_debt is not None and market_cap
                else None
            ),
            # Raw values preserved for the Audit tab
            "_cash": cash_and_cash_equivalents,
            "_total_debt": total_debt,
            "_rev_growth_raw": revenue_growth,
            "_eps_growth_raw": eps_growth,
            "_gross_margin_raw": gross_margin,
        }


def ensure_directories() -> None:
    SECTOR_DIR.mkdir(exist_ok=True)


def sector_root(sector: str) -> Path:
    return SECTOR_DIR / sector


def sector_stock_data_dir(sector: str) -> Path:
    return sector_root(sector) / "stock-data"


def sector_ticker_review_dir(sector: str) -> Path:
    return sector_root(sector) / "ticker-review"


def sector_pillar_tickers_path(sector: str) -> Path:
    return sector_root(sector) / "pillar_tickers.json"


def ensure_sector_directories(sectors: Iterable[str]) -> None:
    for sector in sectors:
        sector_stock_data_dir(sector).mkdir(parents=True, exist_ok=True)
        sector_ticker_review_dir(sector).mkdir(parents=True, exist_ok=True)


def normalize_ticker(value: object) -> str:
    return str(value).strip().upper()


def unique(values: Iterable[str]) -> List[str]:
    seen = set()
    ordered: List[str] = []
    for value in values:
        ticker = normalize_ticker(value)
        if ticker and ticker not in seen:
            seen.add(ticker)
            ordered.append(ticker)
    return ordered


def load_json_file(path: Path) -> dict:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    return data if isinstance(data, dict) else {}


def load_master_config() -> Dict[str, Dict[str, List[str]]]:
    payload = load_json_file(MASTER_JSON_PATH)
    if payload:
        return {
            sector: {
                pillar: unique(tickers)
                for pillar, tickers in sector_payload.items()
                if isinstance(tickers, list)
            }
            for sector, sector_payload in payload.items()
            if isinstance(sector_payload, dict)
        }

    fallback: Dict[str, Dict[str, List[str]]] = {}
    if SECTOR_DIR.exists():
        for child in SECTOR_DIR.iterdir():
            if not child.is_dir():
                continue
            sector_payload = load_json_file(child / "pillar_tickers.json")
            if sector_payload:
                fallback[child.name] = {
                    pillar: unique(tickers)
                    for pillar, tickers in sector_payload.items()
                    if isinstance(tickers, list)
                }
    if fallback:
        return fallback

    return {}


def sync_master_files(master_config: Dict[str, Dict[str, List[str]]]) -> None:
    with MASTER_JSON_PATH.open("w", encoding="utf-8") as handle:
        json.dump(master_config, handle, indent=2)
        handle.write("\n")

    ensure_sector_directories(master_config.keys())
    for sector, sector_config in master_config.items():
        with sector_pillar_tickers_path(sector).open("w", encoding="utf-8") as handle:
            json.dump(sector_config, handle, indent=2)
            handle.write("\n")


def all_pillar_contexts(master_config: Dict[str, Dict[str, List[str]]]) -> List[TickerContext]:
    contexts: List[TickerContext] = []
    for sector, sector_payload in master_config.items():
        for pillar, tickers in sector_payload.items():
            contexts.append(TickerContext(sector=sector, pillar=pillar, tickers=unique(tickers)))
    return contexts


def sheet_name_for_pillar(pillar: str) -> str:
    if pillar in KNOWN_SHEET_NAMES:
        return KNOWN_SHEET_NAMES[pillar]

    cleaned = "".join(char for char in pillar if char.isalnum() or char in [" ", "&", "-"])
    cleaned = cleaned.replace("&", "And").replace("-", " ")
    parts = [part for part in cleaned.split() if part]
    if not parts:
        return pillar[:31]

    abbreviated = []
    for part in parts:
        abbreviated.append(part[:4].title() if len(part) > 4 else part.title())

    return " ".join(abbreviated)[:31]


def pillar_pe_multiple(pillar: str) -> int:
    key = pillar.lower().replace(" ", "_").replace("-", "_")
    return PILLAR_TARGET_PE_MULTIPLES.get(key, PILLAR_TARGET_PE_MULTIPLES["default"])


def _graham_number(eps: Optional[float], bvps: Optional[float]) -> Optional[float]:
    if eps is None or bvps is None or eps <= 0 or bvps <= 0:
        return None
    return math.sqrt(22.5 * eps * bvps)


def compute_investment_score(record: Dict[str, Optional[float]]) -> Optional[float]:
    def clamp(val: float, lo: float, hi: float) -> float:
        return max(lo, min(hi, val))

    def norm(val: Optional[float], lo: float, hi: float) -> Optional[float]:
        if val is None:
            return None
        return clamp((val - lo) / (hi - lo) * 100, 0.0, 100.0)

    def norm_inv(val: Optional[float], lo: float, hi: float) -> Optional[float]:
        if val is None:
            return None
        return clamp((hi - val) / (hi - lo) * 100, 0.0, 100.0)

    # Growth (35%)
    rev_score = norm(record.get("Revenue Growth %"), -50.0, 100.0)       # 20%
    eps_score = norm(record.get("EPS Growth %"), -50.0, 100.0)           # 10%
    backlog_score = norm(record.get("Backlog Growth %"), -50.0, 100.0)   # 5%

    # Financial Quality (25%)
    gm_score = norm(record.get("Gross Margin %"), 0.0, 100.0)            # 15%
    ncr_score = norm(record.get("Net Cash Ratio"), -1.0, 2.0)            # 10%

    # Valuation (25%)
    pe = record.get("P/E Ratio")
    pe_score: Optional[float] = 0.0 if (pe is not None and pe <= 0) else norm_inv(pe, 0.0, 60.0)  # 5%
    ps_score = norm_inv(record.get("P/S Ratio"), 0.0, 30.0)              # 10%
    upside_score = norm(record.get("Upside %"), -30.0, 100.0)  # 10%

    # Entry Timing (15%)
    rsi = record.get("RSI")
    # Score highest at RSI=30 (oversold), linearly to 0 at RSI=80 (overbought)
    rsi_score: Optional[float] = clamp((80.0 - rsi) / 50.0 * 100.0, 0.0, 100.0) if rsi is not None else None  # 10%
    current_price = record.get("Current Price")
    buy_zone = record.get("Buy Zone")
    if current_price is not None and buy_zone is not None and buy_zone > 0:
        # Score 100 at or below buy zone, 0 when 25%+ above it
        ratio = current_price / buy_zone
        bz_score: Optional[float] = clamp((1.25 - ratio) / 0.25 * 100.0, 0.0, 100.0)
    else:
        bz_score = None

    weighted = [
        (rev_score, 0.20),
        (eps_score, 0.10),
        (backlog_score, 0.05),
        (gm_score, 0.15),
        (ncr_score, 0.10),
        (pe_score, 0.05),
        (ps_score, 0.10),
        (upside_score, 0.10),
        (rsi_score, 0.10),
        (bz_score, 0.05),
    ]

    total_weight = sum(w for s, w in weighted if s is not None)
    if total_weight == 0:
        return None
    weighted_sum = sum(s * w for s, w in weighted if s is not None)
    return round(weighted_sum / total_weight, 1)


def fetch_records_for_pillar(
    context: TickerContext,
    client: Optional[FMPClient],
) -> List[Dict[str, Optional[float]]]:
    records: List[Dict[str, Optional[float]]] = []
    for ticker in context.tickers:
        metrics: Dict[str, Optional[float]] = {}
        if client is not None:
            try:
                metrics = client.get_metrics(ticker)
                time.sleep(0.15)
            except Exception as exc:
                print(f"[WARN] Failed to fetch metrics for {ticker}: {exc}")
                metrics = {}

        current_price = metrics.get("Current Price")
        forward_eps = metrics.get("Forward EPS")
        ttm_eps = metrics.get("TTM EPS")
        bvps = metrics.get("Book Value Per Share")

        # Target Price via Forward P/E; fall back to TTM EPS
        eps_used = forward_eps if (forward_eps is not None and forward_eps != 0) else ttm_eps
        eps_source = (
            "Forward" if (forward_eps is not None and forward_eps != 0)
            else ("TTM" if (ttm_eps is not None and ttm_eps != 0) else "None")
        )
        pe_multiple = pillar_pe_multiple(context.pillar)
        if eps_used is not None and eps_used != 0:
            target_price: Optional[float] = eps_used * pe_multiple
        else:
            target_price = None
            print(f"[WARN] Target Price unavailable for {ticker} — missing EPS data")

        # Upside % as a percentage (e.g. 25.0 for 25%)
        if target_price is not None and current_price not in (None, 0):
            upside_pct: Optional[float] = ((target_price - current_price) / current_price) * 100
        else:
            upside_pct = None

        # Graham Number and undervaluation flag
        graham = _graham_number(eps_used, bvps)
        if graham is None:
            graham_undervalued: Optional[bool] = None
        else:
            graham_undervalued = current_price is not None and current_price < graham

        record: Dict[str, object] = {
            "Ticker": ticker,
            "Current Price": current_price,
            "RSI": metrics.get("RSI"),
            "P/E Ratio": metrics.get("P/E Ratio"),
            "P/S Ratio": metrics.get("P/S Ratio"),
            "Market Cap": metrics.get("Market Cap"),
            "Buy Zone": current_price * 0.8 if current_price is not None else None,
            "Target Price": target_price,
            "Graham Undervalued": graham_undervalued,
            "Upside %": upside_pct,
            "Revenue Growth %": metrics.get("Revenue Growth %"),
            "EPS Growth %": metrics.get("EPS Growth %"),
            "Backlog Growth %": None,
            "Gross Margin %": metrics.get("Gross Margin %"),
            "Net Cash Ratio": metrics.get("Net Cash Ratio"),
            # Audit fields — raw API values and intermediates, not written to pillar tabs
            "_audit_forward_eps": forward_eps,
            "_audit_ttm_eps": ttm_eps,
            "_audit_eps_source": eps_source,
            "_audit_eps_used": eps_used,
            "_audit_pillar_pe": pe_multiple,
            "_audit_book_value_per_share": bvps,
            "_audit_graham_number": graham,
            "_audit_cash": metrics.get("_cash"),
            "_audit_total_debt": metrics.get("_total_debt"),
            "_audit_rev_growth_raw": metrics.get("_rev_growth_raw"),
            "_audit_eps_growth_raw": metrics.get("_eps_growth_raw"),
            "_audit_gross_margin_raw": metrics.get("_gross_margin_raw"),
        }
        record["Investment Score"] = compute_investment_score(record)
        records.append(record)

    return records


MASTER_COLUMNS = ["Pillar"] + COLUMNS


def write_master_sheet(
    workbook: Workbook,
    records_by_context: List[Tuple[TickerContext, List[Dict[str, Optional[float]]]]],
) -> None:
    all_rows: List[Dict[str, object]] = []
    for context, records in records_by_context:
        for record in records:
            row = dict(record)
            row["Pillar"] = context.pillar
            all_rows.append(row)

    all_rows.sort(
        key=lambda r: (r.get("Investment Score") is None, -(r.get("Investment Score") or 0))
    )

    if "Investment Master" in workbook.sheetnames:
        sheet = workbook["Investment Master"]
    else:
        sheet = workbook.create_sheet(title="Investment Master", index=1)

    for col_idx, header in enumerate(MASTER_COLUMNS, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    for row_idx, row in enumerate(all_rows, start=2):
        for col_idx, header in enumerate(MASTER_COLUMNS, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=row.get(header))


AUDIT_COLUMNS = [
    "Pillar", "Ticker", "Current Price",
    "Forward EPS", "TTM EPS", "EPS Source", "EPS Used", "Pillar P/E Multiple",
    "Book Value Per Share", "Cash", "Total Debt", "Market Cap",
    "Revenue Growth (raw)", "EPS Growth (raw)", "Gross Margin (raw)",
    "P/E Ratio", "P/S Ratio", "RSI",
    "Buy Zone", "Target Price", "Graham Number", "Graham Undervalued",
    "Upside %", "Net Cash Ratio", "Revenue Growth %", "EPS Growth %", "Gross Margin %",
    "Investment Score",
]


def write_audit_sheet(
    workbook: Workbook,
    records_by_context: List[Tuple[TickerContext, List[Dict[str, Optional[float]]]]],
) -> None:
    if "Audit" in workbook.sheetnames:
        sheet = workbook["Audit"]
    else:
        sheet = workbook.create_sheet(title="Audit")

    for col_idx, header in enumerate(AUDIT_COLUMNS, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    row_idx = 2
    for context, records in records_by_context:
        for record in records:
            audit_row = {
                "Pillar": context.pillar,
                "Ticker": record.get("Ticker"),
                "Current Price": record.get("Current Price"),
                "Forward EPS": record.get("_audit_forward_eps"),
                "TTM EPS": record.get("_audit_ttm_eps"),
                "EPS Source": record.get("_audit_eps_source"),
                "EPS Used": record.get("_audit_eps_used"),
                "Pillar P/E Multiple": record.get("_audit_pillar_pe"),
                "Book Value Per Share": record.get("_audit_book_value_per_share"),
                "Cash": record.get("_audit_cash"),
                "Total Debt": record.get("_audit_total_debt"),
                "Market Cap": record.get("Market Cap"),
                "Revenue Growth (raw)": record.get("_audit_rev_growth_raw"),
                "EPS Growth (raw)": record.get("_audit_eps_growth_raw"),
                "Gross Margin (raw)": record.get("_audit_gross_margin_raw"),
                "P/E Ratio": record.get("P/E Ratio"),
                "P/S Ratio": record.get("P/S Ratio"),
                "RSI": record.get("RSI"),
                "Buy Zone": record.get("Buy Zone"),
                "Target Price": record.get("Target Price"),
                "Graham Number": record.get("_audit_graham_number"),
                "Graham Undervalued": record.get("Graham Undervalued"),
                "Upside %": record.get("Upside %"),
                "Net Cash Ratio": record.get("Net Cash Ratio"),
                "Revenue Growth %": record.get("Revenue Growth %"),
                "EPS Growth %": record.get("EPS Growth %"),
                "Gross Margin %": record.get("Gross Margin %"),
                "Investment Score": record.get("Investment Score"),
            }
            for col_idx, header in enumerate(AUDIT_COLUMNS, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=audit_row.get(header))
            row_idx += 1


def write_formula_guide_sheet(workbook: Workbook) -> None:
    if "Formula Guide" in workbook.sheetnames:
        sheet = workbook["Formula Guide"]
    else:
        sheet = workbook.create_sheet(title="Formula Guide")

    headers = ["Column", "Formula", "How to Verify", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    formulas = [
        (
            "Buy Zone",
            "Current Price × 0.80",
            "Take Current Price from the Audit tab and multiply by 0.80. Result should match Buy Zone.",
            "Represents a 20% discount to today's price — a target entry level if the stock pulls back.",
        ),
        (
            "Target Price",
            "EPS Used × Pillar P/E Multiple",
            "Multiply EPS Used by Pillar P/E Multiple (both in Audit tab). Result should match Target Price.",
            "EPS Source column shows whether Forward EPS or TTM EPS was used. Pillar P/E multiples are defined in PILLAR_TARGET_PE_MULTIPLES in data_engine.py.",
        ),
        (
            "Graham Number",
            "√(22.5 × EPS Used × Book Value Per Share)",
            "Multiply 22.5 × EPS Used × Book Value Per Share (all in Audit tab), then take the square root.",
            "Only valid when both EPS and Book Value Per Share are positive. A classic Benjamin Graham intrinsic value estimate.",
        ),
        (
            "Graham Undervalued",
            "Current Price < Graham Number",
            "Compare Current Price to Graham Number in the Audit tab. TRUE = stock trades below Graham's estimated value.",
            "Informational only — does not affect Investment Score. None when EPS or Book Value data is unavailable.",
        ),
        (
            "Upside %",
            "((Target Price − Current Price) / Current Price) × 100",
            "Subtract Current Price from Target Price, divide by Current Price, multiply by 100.",
            "Expressed as a percentage (e.g. 25.0 = 25% upside). Negative means the model sees downside at the current price.",
        ),
        (
            "Net Cash Ratio",
            "(Cash − Total Debt) / Market Cap",
            "Subtract Total Debt from Cash, then divide by Market Cap (all in Audit tab).",
            "Positive = net cash position. Negative = net debt. Sourced from FMP balance-sheet-statement.",
        ),
        (
            "Revenue Growth %",
            "Revenue Growth (raw) × 100",
            "Take Revenue Growth (raw) from Audit tab and multiply by 100.",
            "FMP returns growth as a decimal (e.g. 0.185). Multiplying by 100 gives 18.5%.",
        ),
        (
            "EPS Growth %",
            "EPS Growth (raw) × 100",
            "Take EPS Growth (raw) from Audit tab and multiply by 100.",
            "FMP returns growth as a decimal. Multiplying by 100 gives the percentage.",
        ),
        (
            "Gross Margin %",
            "Gross Margin (raw) × 100",
            "Take Gross Margin (raw) from Audit tab and multiply by 100.",
            "Sourced from FMP ratios-ttm-bulk field grossProfitMarginTTM. Multiplying by 100 gives the percentage.",
        ),
        (
            "Investment Score",
            "Weighted composite of 10 metrics, each normalized 0–100, then combined by weight.",
            "See the component weight table below. Each metric can be traced back to the Audit tab.",
            "Missing metrics are excluded and the remaining weights are rescaled proportionally.",
        ),
    ]

    for row_idx, (col, formula, how_to, notes) in enumerate(formulas, start=2):
        sheet.cell(row=row_idx, column=1, value=col)
        sheet.cell(row=row_idx, column=2, value=formula)
        sheet.cell(row=row_idx, column=3, value=how_to)
        sheet.cell(row=row_idx, column=4, value=notes)

    score_start = len(formulas) + 4
    sheet.cell(row=score_start, column=1, value="Investment Score — Component Weights and Normalization")
    score_start += 1

    score_headers = ["Category", "Metric", "Weight", "Scoring Direction", "Normalization Range"]
    for col_idx, h in enumerate(score_headers, start=1):
        sheet.cell(row=score_start, column=col_idx, value=h)
    score_start += 1

    score_components = [
        ("Growth (35%)", "Revenue Growth %", "20%", "Higher is better", "-50% to 100%"),
        ("Growth (35%)", "EPS Growth %", "10%", "Higher is better", "-50% to 100%"),
        ("Growth (35%)", "Backlog Growth %", "5%", "Higher is better", "-50% to 100%"),
        ("Financial Quality (25%)", "Gross Margin %", "15%", "Higher is better", "0% to 100%"),
        ("Financial Quality (25%)", "Net Cash Ratio", "10%", "Higher is better", "-1.0 to 2.0"),
        ("Valuation (25%)", "P/S Ratio", "10%", "Lower is better", "0 to 30 (capped; above 30 = score 0)"),
        ("Valuation (25%)", "Upside %", "10%", "Higher is better", "-30% to 100%"),
        ("Valuation (25%)", "P/E Ratio", "5%", "Lower is better", "0 to 60 (capped; above 60 or negative = score 0)"),
        ("Entry Timing (15%)", "RSI", "10%", "Lower is better (oversold favored)", "Score 100 at RSI=30, score 0 at RSI=80+"),
        ("Entry Timing (15%)", "Buy Zone Proximity", "5%", "Closer to Buy Zone = higher score", "Score 100 at/below Buy Zone, score 0 when 25%+ above it"),
    ]

    for comp in score_components:
        for col_idx, val in enumerate(comp, start=1):
            sheet.cell(row=score_start, column=col_idx, value=val)
        score_start += 1


def write_records_to_sheet(workbook: Workbook, sheet_name: str, records: List[Dict[str, Optional[float]]]) -> None:
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)

    for column_index, header in enumerate(COLUMNS, start=1):
        sheet.cell(row=1, column=column_index, value=header)

    for row_index, record in enumerate(records, start=2):
        for column_index, header in enumerate(COLUMNS, start=1):
            sheet.cell(row=row_index, column=column_index, value=record.get(header))


def format_metric(value: Optional[float], suffix: str = "") -> str:
    if value is None:
        return "n/a"
    if isinstance(value, float):
        return f"{value:.2f}{suffix}"
    return f"{value}{suffix}"


def build_summary(records_by_context: List[Tuple[TickerContext, List[Dict[str, Optional[float]]]]]) -> Dict[str, object]:
    all_records: List[Dict[str, object]] = []
    pillar_counts: Dict[str, int] = {}

    for context, records in records_by_context:
        pillar_counts[context.pillar] = len(records)
        for record in records:
            row = dict(record)
            row["Sector"] = context.sector
            row["Pillar"] = context.pillar
            all_records.append(row)

    if not all_records:
        return {"total_tickers": 0, "top_candidates": [], "pillar_counts": {}, "all_records": []}

    data_frame = pd.DataFrame(all_records)
    data_frame = data_frame.sort_values(by=["Upside %", "Ticker"], ascending=[False, True], na_position="last")
    top_candidates = data_frame.head(5)[
        ["Sector", "Pillar", "Ticker", "Upside %", "Revenue Growth %", "EPS Growth %", "Gross Margin %", "Net Cash Ratio"]
    ].to_dict("records")

    return {
        "total_tickers": int(len(data_frame)),
        "pillar_counts": pillar_counts,
        "top_candidates": top_candidates,
        "all_records": all_records,
    }


# Claude ticker review — commented out to avoid API usage costs.
# Uncomment this function and the call site in run() to re-enable.
#
# def generate_claude_ticker_review(summary: Dict[str, object], date_stamp: str, sector: str) -> Optional[str]:
#     api_key = os.getenv("ANTHROPIC_API_KEY")
#     if not api_key:
#         return None
#
#     try:
#         import anthropic
#     except Exception:
#         return None
#
#     client = anthropic.Anthropic(api_key=api_key)
#     prompt = (
#         f"Sector: {sector}. Data date: {date_stamp}. "
#         f"Structured data: {json.dumps(summary.get('all_records', []), default=str)}. "
#         "Given this data, use deep research to assess similar stocks and please provide 5 (Max) tickers to review."
#     )
#     response = client.messages.create(
#         model="claude-sonnet-4-6",
#         max_tokens=512,
#         system="You are a buy-side investment research assistant.",
#         messages=[{"role": "user", "content": prompt}],
#     )
#     return next((block.text for block in response.content if block.type == "text"), None)


def write_text(path: Path, content: str) -> None:
    path.write_text(content.rstrip() + "\n", encoding="utf-8")


def run(api_key: Optional[str]) -> Dict[str, object]:
    ensure_directories()
    master_config = load_master_config()
    sync_master_files(master_config)
    ensure_sector_directories(master_config.keys())

    client = FMPClient(api_key) if api_key else None
    date_stamp = datetime.now().strftime(DATE_FORMAT)

    contexts_by_sector: Dict[str, List[TickerContext]] = {}
    for context in all_pillar_contexts(master_config):
        contexts_by_sector.setdefault(context.sector, []).append(context)

    outputs: Dict[str, object] = {"date": date_stamp, "sectors": {}}

    for sector, sector_contexts in contexts_by_sector.items():
        workbook = Workbook()
        workbook.remove(workbook.active)

        records_by_context: List[Tuple[TickerContext, List[Dict[str, Optional[float]]]]] = []

        for context in sector_contexts:
            sheet_name = sheet_name_for_pillar(context.pillar)
            records = fetch_records_for_pillar(context, client)
            write_records_to_sheet(workbook, sheet_name, records)
            records_by_context.append((context, records))

        write_master_sheet(workbook, records_by_context)
        write_audit_sheet(workbook, records_by_context)
        write_formula_guide_sheet(workbook)

        daily_workbook_path = sector_stock_data_dir(sector) / f"{DAILY_WORKBOOK_PREFIX}_{date_stamp}.xlsx"
        workbook.save(daily_workbook_path)
        workbook.close()

        # Claude ticker review disabled — uncomment to re-enable once generate_claude_ticker_review is restored.
        # summary = build_summary(records_by_context)
        # ticker_review_text = generate_claude_ticker_review(summary, date_stamp, sector)
        # if ticker_review_text:
        #     ticker_review_path = sector_ticker_review_dir(sector) / f"{TICKER_REVIEW_PREFIX}_{date_stamp}.txt"
        #     write_text(ticker_review_path, ticker_review_text)

        outputs["sectors"][sector] = {
            "daily_workbook": str(daily_workbook_path),
        }

    return outputs


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run the investment engine daily refresh")
    parser.add_argument("--api-key", default=None, help="Override FMP API key")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    api_key = args.api_key or os.getenv("FMP") or os.getenv("FMP_API_KEY")
    outputs = run(api_key=api_key)
    print(json.dumps(outputs, indent=2))


if __name__ == "__main__":
    main()
